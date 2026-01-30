#!/usr/bin/env python

from datetime import datetime
import io
import re
import sys
import urllib3
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

urllib3.disable_warnings(urllib3.exceptions.NotOpenSSLWarning)

# Import all version checkers
from src.checkers.github import get_github_latest_version, get_github_latest_tag
from src.checkers.home_assistant import get_home_assistant_version
from src.checkers.esphome import get_esphome_version
from src.checkers.konnected import get_konnected_version, get_konnected_current_version
from src.checkers.airgradient import (
    get_airgradient_version,
    get_airgradient_current_version,
)
from src.checkers.opnsense import get_opnsense_version
from src.checkers.k3s import get_k3s_current_version
from src.checkers.zigbee2mqtt import get_zigbee2mqtt_version
from src.checkers.kopia import get_kopia_version
from src.checkers.kubectl import (
    get_telegraf_version,
    get_mosquitto_version,
    get_victoriametrics_version,
    get_calico_version,
    get_metallb_version,
    get_alertmanager_version,
    get_fluentbit_version,
    get_mongodb_version,
    get_opensearch_version,
    get_pgadmin_version,
    get_unpoller_version,
    get_certmanager_version,
    get_postfix_version,
    get_minio_kubectl_version,
)
from src.checkers.cnpg import get_cnpg_version, get_cnpg_postgres_latest_version
from src.checkers.server_status import check_server_status
from src.checkers.proxmox import get_proxmox_version, get_proxmox_latest_version
from src.checkers.tailscale import check_tailscale_versions
from src.checkers.traefik import get_traefik_version
from src.checkers.graylog import (
    get_graylog_current_version,
    get_graylog_latest_version_from_repo,
    get_postgresql_latest_version_from_ghcr,
)
from src.checkers.grafana import get_grafana_version
from src.checkers.mongodb import get_mongodb_latest_version
from src.checkers.unifi_protect import (
    get_unifi_protect_version,
    get_unifi_protect_latest_version,
)
from src.checkers.unifi_network import (
    get_unifi_network_version,
    get_unifi_network_latest_version,
)
from src.checkers.unifi_os import get_unifi_os_nvr_latest_version, get_unifi_os_version
from src.checkers.samba import get_samba_version, get_latest_samba_version
from src.checkers.syncthing import check_syncthing_current_version
from src.checkers.awx import check_awx_current_version
from src.checkers.graylog_compat import get_opensearch_compatible_version
from src.checkers.postfix import get_postfix_latest_version_from_dockerhub
from src.checkers.dockerhub import (
    get_dockerhub_latest_version,
    get_dockerhub_latest_beta,
)
from src.checkers.n8n import get_n8n_version_api, get_n8n_version_kubectl
from src.checkers.wyoming import (
    get_wyoming_openwakeword_version,
    get_wyoming_piper_version,
    get_wyoming_whisper_version,
    get_wyoming_satellite_version,
)
from src.checkers.ollama import get_ollama_version
from src.checkers.docker import get_docker_version
from src.checkers.portainer import get_portainer_version
from src.checkers.open_webui import get_open_webui_version
from src.checkers.uptime_kuma import (
    get_uptime_kuma_version as get_uptime_kuma_api_version,
)
import config


class VersionManager:
    # Constants
    DEFAULT_EXCEL_PATH = config.EXCEL_FILE_PATH
    SHEET_NAME = "Sheet1"

    STATUS_ICONS = {
        "Up to Date": "✅",
        "Update Available": "⚠️ ",
        "Latest Available": "📋",
        "Current Version": "📌",
        "Unknown": "❓",
    }

    # Expected column names (order-independent)
    EXPECTED_COLUMNS = [
        "Name",
        "Enabled",
        "Instance",
        "Type",
        "Category",
        "Target",
        "GitHub",
        "DockerHub",
        "Current_Version",
        "Latest_Version",
        "Status",
        "Last_Checked",
        "Check_Current",
        "Check_Latest",
        "Key",
    ]

    def __init__(self, excel_path=None):
        self.excel_path = excel_path or self.DEFAULT_EXCEL_PATH
        self.workbook = None
        self.worksheet = None
        self.columns = {}  # Dynamic column mapping
        self._lock = threading.Lock()  # Thread-safe Excel updates
        self.load_workbook()

    def load_workbook(self):
        """Load Excel workbook using openpyxl and map columns dynamically"""
        try:
            self.workbook = load_workbook(self.excel_path)
            self.worksheet = self.workbook[self.SHEET_NAME]

            # Read header row and create dynamic column mapping
            self._map_columns()

            # Count applications (excluding header row)
            app_count = self.worksheet.max_row - 1 if self.worksheet.max_row > 1 else 0
            print(f"Loaded {app_count} applications from Excel")
        except FileNotFoundError:
            print(f"Excel file not found: {self.excel_path}")
            self.workbook = None
            self.worksheet = None
            self.columns = {}
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            self.workbook = None
            self.worksheet = None
            self.columns = {}

    def _map_columns(self):
        """Create dynamic column mapping by reading header row"""
        self.columns = {}
        if not self.worksheet:
            return

        # Read header row (row 1) and map column names to letters
        for col_num in range(1, self.worksheet.max_column + 1):
            col_letter = chr(ord("A") + col_num - 1)
            header_cell = self.worksheet[f"{col_letter}1"]

            if header_cell.value:
                column_name = str(header_cell.value).strip()
                self.columns[column_name] = col_letter

        # Verify all expected columns are present
        missing_columns = []
        for expected_col in self.EXPECTED_COLUMNS:
            if expected_col not in self.columns:
                missing_columns.append(expected_col)

        if missing_columns:
            print(f"Warning: Missing columns in Excel file: {missing_columns}")
            print(f"Found columns: {list(self.columns.keys())}")

    def save_workbook(self):
        """Save workbook preserving all formatting"""
        if self.workbook is None:
            print("No workbook to save")
            return
        try:
            self.workbook.save(self.excel_path)
            print("Excel file updated successfully (formatting preserved)")
        except Exception as e:
            print(f"Error saving Excel file: {e}")

    def get_row_data(self, row_num):
        """Get all data for a specific row"""
        if self.worksheet is None or row_num < 2:
            return None

        data = {}
        for col_name, col_letter in self.columns.items():
            cell = self.worksheet[f"{col_letter}{row_num}"]
            data[col_name] = cell.value if cell.value is not None else ""

        return data

    def update_row_data(self, row_num, updates):
        """Update specific columns in a row (thread-safe)"""
        if self.worksheet is None:
            return

        with self._lock:
            for col_name, value in updates.items():
                if col_name in self.columns:
                    col_letter = self.columns[col_name]
                    self.worksheet[f"{col_letter}{row_num}"] = value

    def find_application_row(self, app_name, instance="prod"):
        """Find the row number for a specific application and instance"""
        if (
            self.worksheet is None
            or "Name" not in self.columns
            or "Instance" not in self.columns
        ):
            return None

        for row_num in range(2, self.worksheet.max_row + 1):
            name_cell = self.worksheet[f"{self.columns['Name']}{row_num}"]
            instance_cell = self.worksheet[f"{self.columns['Instance']}{row_num}"]

            if (
                name_cell.value
                and name_cell.value.lower() == app_name.lower()
                and instance_cell.value
                and instance_cell.value.lower() == instance.lower()
            ):
                return row_num
        return None

    def find_application_rows_by_name(self, app_name):
        """Find all row numbers for a specific application name (all instances)

        Args:
            app_name: Application name to search for (case-insensitive)

        Returns:
            List of row numbers matching the application name
        """
        if self.worksheet is None or "Name" not in self.columns:
            return []

        matching_rows = []
        for row_num in range(2, self.worksheet.max_row + 1):
            # Skip disabled applications
            if "Enabled" in self.columns:
                enabled_cell = self.worksheet[f"{self.columns['Enabled']}{row_num}"]
                if enabled_cell.value is not True:
                    continue
            name_cell = self.worksheet[f"{self.columns['Name']}{row_num}"]
            if name_cell.value and name_cell.value.lower() == app_name.lower():
                matching_rows.append(row_num)

        return matching_rows

    def get_all_application_names(self):
        """Get a sorted list of unique application names

        Returns:
            Sorted list of unique application names
        """
        if self.worksheet is None or "Name" not in self.columns:
            return []

        names = set()
        for row_num in range(2, self.worksheet.max_row + 1):
            # Skip disabled applications
            if "Enabled" in self.columns:
                enabled_cell = self.worksheet[f"{self.columns['Enabled']}{row_num}"]
                if enabled_cell.value is not True:
                    continue
            name_cell = self.worksheet[f"{self.columns['Name']}{row_num}"]
            if name_cell.value:
                names.add(name_cell.value)

        return sorted(names)

    def _get_dockerhub_version_for_app(
        self, app_name, dockerhub_repo, version_pin=None
    ):
        """Helper method to get Docker Hub version for specific applications"""
        if version_pin == "beta":
            return get_dockerhub_latest_beta(dockerhub_repo)
        if app_name == "mongodb":
            return get_mongodb_latest_version()
        elif app_name == "graylog":
            if "graylog" in dockerhub_repo.lower():
                return get_graylog_latest_version_from_repo(dockerhub_repo)
            elif "postgres" in dockerhub_repo.lower():
                return get_postgresql_latest_version_from_ghcr(dockerhub_repo)
            elif "postfix" in dockerhub_repo.lower():
                return get_postfix_latest_version_from_dockerhub(dockerhub_repo)
            else:
                return get_dockerhub_latest_version(dockerhub_repo)
        elif app_name == "minio":
            # MinIO uses RELEASE.YYYY-MM-DDTHH-MM-SSZ format
            return get_dockerhub_latest_version(
                dockerhub_repo,
                version_pattern=r"^RELEASE\.(\d{4}-\d{2}-\d{2}T\d{2}-\d{2}-\d{2}Z)(?:-[a-z0-9]+)?$",
                exclude_tags=["latest"],
            )
        else:
            return get_dockerhub_latest_version(dockerhub_repo)

    def _get_github_version_for_app(self, app_name, github_repo, check_latest):
        """Helper method to get GitHub version for specific applications"""
        if app_name == "mongodb" and github_repo == "mongodb/mongo":
            # Only use hardcoded MongoDB function for database instances
            return get_mongodb_latest_version()
        elif check_latest == "github_release":
            return get_github_latest_version(github_repo)
        elif check_latest == "github_tag":
            return get_github_latest_tag(github_repo)
        return None

    def get_latest_version(
        self, app_name, check_latest, github_repo, dockerhub_repo, version_pin=None
    ):
        """Get latest version based on check_latest method"""
        latest_version = None

        # Get latest version based on check_latest method
        if check_latest == "github_release" or check_latest == "github_tag":
            if github_repo and github_repo.strip():
                # Special case for Konnected - use YAML project_version instead of GitHub tags
                if app_name == "konnected":
                    latest_version = get_konnected_version("latest", None, github_repo)
                else:
                    latest_version = self._get_github_version_for_app(
                        app_name, github_repo, check_latest
                    )
        elif check_latest == "docker_hub":
            if dockerhub_repo and dockerhub_repo.strip():
                # Special case for CNPG PostgreSQL - use catalog instead of docker hub
                if app_name == "cnpg" and "postgres-containers" in dockerhub_repo:
                    latest_version = get_cnpg_postgres_latest_version()
                else:
                    latest_version = self._get_dockerhub_version_for_app(
                        app_name, dockerhub_repo, version_pin
                    )
        elif check_latest == "proxmox":
            latest_version = get_proxmox_latest_version(include_ceph=True)
        elif check_latest == "unifi_protect_rss":
            # Special case for UniFi Protect using RSS feed
            if app_name == "ui-protect":
                latest_version = get_unifi_protect_latest_version()
        elif check_latest == "unifi_network_rss":
            # Special case for UniFi Network using RSS feed
            if app_name == "ui-network":
                latest_version = get_unifi_network_latest_version()
        elif check_latest == "unifi_os_nvr_rss":
            # Special case for UniFi OS NVR/UNVR using RSS feed
            latest_version = get_unifi_os_nvr_latest_version()
        elif check_latest == "graylog_compat":
            latest_version = get_opensearch_compatible_version()
        elif check_latest == "helm_chart":
            # Special case for MongoDB operator using Helm chart version
            if app_name == "mongodb":
                from src.checkers.utils import get_helm_chart_version

                latest_version = get_helm_chart_version(
                    "mongodb/helm-charts", "community-operator", "operator.version"
                )
            # Special case for Fluent Bit using Helm chart appVersion
            elif app_name == "fluent-bit":
                from src.checkers.utils import get_helm_chart_app_version

                latest_version = get_helm_chart_app_version(
                    "fluent/helm-charts", "fluent-bit"
                )
        # ssh_apt method - latest version will be populated during ssh current check

        return latest_version

    def get_current_version(self, app_data):
        """Get current version based on check_current method"""
        app_name = app_data.get("Name", "")
        instance = app_data.get("Instance", "prod")
        check_current = app_data.get("Check_Current", "")
        check_latest = app_data.get("Check_Latest", "")
        url = app_data.get("Target", "")
        github_repo = app_data.get("GitHub", "")
        context = app_data.get("Context", "") or None
        namespace = app_data.get("Namespace", "") or None

        current_version = None
        latest_version = None
        firmware_update_available = False

        # Get current version based on check_current method
        if check_current == "api":
            if app_name == "home-assistant":
                if url:
                    current_version = get_home_assistant_version(instance, url)
            elif app_name == "esphome":
                if url:
                    current_version = get_esphome_version(url)
            elif app_name == "konnected":
                # Check device API for current version
                current_version = get_konnected_current_version(instance, url)
            elif app_name == "airgradient":
                # Check device API for current version with encryption key
                encryption_key = app_data.get("Key", "")
                current_version = get_airgradient_current_version(
                    instance, url, encryption_key
                )
            elif app_name == "traefik":
                current_version = get_traefik_version(instance, url)
            elif app_name == "opnsense":
                result = get_opnsense_version(instance, url)
                if isinstance(result, dict):
                    current_version = result.get("current_version")
                    firmware_update_available = result.get(
                        "firmware_update_available", False
                    )

                    # Use the full_version as the latest version for OPNsense
                    if result.get("full_version"):
                        latest_version = result["full_version"]
            elif app_name == "proxmox":
                if url:
                    current_version = get_proxmox_version(instance, url)
            elif app_name == "tailscale":
                print("  Checking all Tailscale devices...")
                device_results = check_tailscale_versions(
                    api_key=config.TAILSCALE_API_KEY, tailnet=config.TAILSCALE_TAILNET
                )

                if device_results and device_results["total_devices"] > 0:
                    # Use counts from API update availability
                    devices_needing_updates = device_results["devices_needing_updates"]
                    devices_up_to_date = device_results["devices_up_to_date"]

                    # Current version shows count of devices needing updates
                    current_version = f"{devices_needing_updates} need updates"

                    # Latest version shows count of devices up to date
                    latest_version = f"{devices_up_to_date} up-to-date"
            elif app_name == "graylog":
                if url:
                    current_version = get_graylog_current_version(instance, url)
            elif app_name == "ui-protect":
                if url:
                    current_version = get_unifi_protect_version(instance, url)
            elif app_name == "ui-network":
                if url:
                    current_version = get_unifi_network_version(instance, url)
            elif app_name == "awx":
                if url:
                    current_version = check_awx_current_version(instance, url)
            elif app_name == "syncthing":
                if url:
                    current_version = check_syncthing_current_version(instance, url)
            elif app_name == "n8n":
                if url:
                    current_version = get_n8n_version_api(instance, url)
            elif app_name == "ollama":
                if url:
                    current_version = get_ollama_version(instance, url)
            elif app_name == "portainer":
                if url:
                    current_version = get_portainer_version(instance, url)
            elif app_name == "open-webui":
                if url:
                    current_version = get_open_webui_version(instance, url)
            elif app_name == "uptime-kuma":
                current_version = get_uptime_kuma_api_version(instance, url)
            # Add more API-based applications here as needed

        elif check_current == "ssh":
            if check_latest == "ssh_apt":
                # SSH with kernel checking - this will set both current and latest versions
                kernel_result = check_server_status(instance, url)
                if isinstance(kernel_result, dict):
                    current_version = kernel_result.get("current_version")
                    latest_version = kernel_result.get("latest_version")
            else:
                # SSH for current version only - handle specific applications
                if app_name == "unifi-os":
                    if url:
                        current_version = get_unifi_os_version(instance, url)

        elif check_current == "kubectl":
            if app_name == "telegraf":
                current_version = get_telegraf_version(instance, context=context, namespace=namespace)
            elif app_name == "victoriametrics":
                current_version = get_victoriametrics_version(instance, context=context, namespace=namespace)
            elif app_name == "mosquitto":
                current_version = get_mosquitto_version(instance, context=context, namespace=namespace)
            elif app_name == "calico":
                current_version = get_calico_version(instance, context=context, namespace=namespace)
            elif app_name == "metallb":
                current_version = get_metallb_version(instance, context=context, namespace=namespace)
            elif app_name == "alertmanager":
                current_version = get_alertmanager_version(instance, context=context, namespace=namespace)
            elif app_name == "fluent-bit":
                current_version = get_fluentbit_version(instance, context=context, namespace=namespace)
            elif app_name == "mongodb":
                current_version = get_mongodb_version(instance, context=context, namespace=namespace)
            elif app_name == "opensearch":
                current_version = get_opensearch_version(instance, context=context, namespace=namespace)
            elif app_name == "cnpg":
                current_version = get_cnpg_version(instance, context=context, namespace=namespace)
            elif app_name == "pgadmin":
                current_version = get_pgadmin_version(instance, context=context, namespace=namespace)
            elif app_name == "grafana":
                current_version = get_grafana_version(instance, context=context, namespace=namespace)
            elif app_name == "unpoller":
                current_version = get_unpoller_version(instance, context=context, namespace=namespace)
            elif app_name == "cert-manager":
                current_version = get_certmanager_version(instance, context=context, namespace=namespace)
            elif app_name == "k3s":
                current_version = get_k3s_current_version(instance, context=context)
            elif app_name == "postfix":
                current_version = get_postfix_version(instance, context=context, namespace=namespace)
            elif app_name == "minio":
                current_version = get_minio_kubectl_version(instance, context=context, namespace=namespace)
            elif app_name == "n8n":
                current_version = get_n8n_version_kubectl(instance, context=context, namespace=namespace)
            elif app_name == "rhasspy" and instance == "wyoming-openwakeword":
                current_version = get_wyoming_openwakeword_version(instance, url)
            elif app_name == "rhasspy" and instance == "wyoming-piper":
                current_version = get_wyoming_piper_version(instance, url)
            elif app_name == "rhasspy" and instance == "wyoming-whisper":
                current_version = get_wyoming_whisper_version(instance, url)

        elif check_current == "mqtt":
            if app_name == "zigbee2mqtt":
                current_version = get_zigbee2mqtt_version(instance)

        elif check_current == "command":
            if app_name == "kopia":
                current_version = get_kopia_version(instance, url)
            elif app_name == "samba":
                current_version = get_samba_version(instance, url)
                if check_latest == "ssh_apt":
                    latest_version = get_latest_samba_version(instance)
            elif app_name == "docker":
                current_version = get_docker_version(instance, url)
            elif app_name == "wyoming-satellite":
                current_version = get_wyoming_satellite_version(instance, url)

        return current_version, latest_version, firmware_update_available

    def check_single_application(self, row_num):
        """Check version for a single application by row number"""
        if self.worksheet is None:
            print("No worksheet loaded")
            return

        app_data = self.get_row_data(row_num)
        if not app_data:
            return

        app_name = app_data.get("Name", "")
        instance = app_data.get("Instance", "prod")
        check_current = app_data.get("Check_Current", "")
        check_latest = app_data.get("Check_Latest", "")
        github_repo = app_data.get("GitHub", "")
        dockerhub_repo = app_data.get("DockerHub", "")

        print(f"Checking {app_name} ({instance})...")

        # Special handling for apps that need current version for smart latest version logic
        if app_name == "unifi-os" and check_latest == "unifi_os_nvr_rss":
            # Get current version first
            current_version, ssh_latest_version, firmware_update_available = (
                self.get_current_version(app_data)
            )
            # Then get latest version with current version context
            latest_version = get_unifi_os_nvr_latest_version(current_version)
        elif check_latest == "proxmox":
            # Get current version first for Proxmox to enable smart version comparison
            current_version, ssh_latest_version, firmware_update_available = (
                self.get_current_version(app_data)
            )
            # Then get latest version with current version context
            latest_version = get_proxmox_latest_version(
                include_ceph=True, current_version=current_version
            )
        else:
            # Standard flow: get latest version first
            version_pin = app_data.get("Version_Pin", "")
            latest_version = self.get_latest_version(
                app_name, check_latest, github_repo, dockerhub_repo, version_pin
            )

            # Get current version (and possibly additional latest version from SSH methods)
            current_version, ssh_latest_version, firmware_update_available = (
                self.get_current_version(app_data)
            )

        # For SSH-based methods, use the latest version from the SSH check if available
        if ssh_latest_version:
            latest_version = ssh_latest_version

        # Update timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Prepare updates
        updates = {"Last_Checked": timestamp}

        updates["Current_Version"] = current_version if current_version else ""
        updates["Latest_Version"] = latest_version if latest_version else ""

        # Update notes if firmware update is available (OPNsense specific)
        if firmware_update_available:
            updates["Notes"] = "Firmware update available"

        # Determine status
        status = "Unknown"
        if current_version and latest_version:
            # Special handling for Tailscale multi-device checking
            if check_current == "api" and app_name == "tailscale":
                # For Tailscale, current_version contains devices needing updates count
                if "0 need updates" in current_version:
                    status = "Up to Date"
                else:
                    status = "Update Available"
            # Special handling for SSH systems with ssh_apt method
            elif check_current == "ssh" and check_latest == "ssh_apt":
                # For SSH systems, status is based on whether apt found updates
                if latest_version == "update available":
                    status = "Update Available"
                elif latest_version == "No updates":
                    status = "Up to Date"
                else:
                    status = "Up to Date"
            else:
                # Standard version comparison for other applications
                # For Kopia, extract just the version number for comparison
                current_clean = current_version
                if "build:" in current_version:
                    current_clean = current_version.split("build:")[0].strip()

                # Remove 'v' prefix if present in latest_version
                latest_clean = latest_version
                if latest_clean.startswith("v"):
                    latest_clean = latest_clean[1:]

                if current_clean == latest_clean:
                    status = "Up to Date"
                else:
                    status = "Update Available"
        elif latest_version and not current_version:
            status = "Latest Available"
        elif current_version and not latest_version:
            status = "Current Version"

        updates["Status"] = status

        # Update the Excel row
        self.update_row_data(row_num, updates)

        # Display results
        current_display = current_version if current_version else "N/A"
        latest_display = latest_version if latest_version else "N/A"
        icon = self.STATUS_ICONS.get(status, "")

        print(f"  Current: {current_display}")
        print(f"  Latest: {latest_display}")
        print(f"  Status: {icon} {status}")
        print()

    def check_all_applications(self, max_workers=10):
        """Check versions for all applications concurrently

        Args:
            max_workers: Maximum number of concurrent threads (default: 10)
        """
        if self.worksheet is None:
            print("No worksheet loaded")
            return

        print(
            f"Starting concurrent version check for all applications (max {max_workers} workers)..."
        )
        print("=" * 50)

        # Collect row numbers for enabled applications only
        row_nums = []
        skipped = 0
        for row_num in range(2, self.worksheet.max_row + 1):
            # Check if Enabled column exists and is True
            if "Enabled" in self.columns:
                enabled_cell = self.worksheet[f"{self.columns['Enabled']}{row_num}"]
                if enabled_cell.value is True:
                    row_nums.append(row_num)
                else:
                    skipped += 1
            else:
                # If no Enabled column, check all applications (backward compatibility)
                row_nums.append(row_num)

        total_apps = len(row_nums)
        if skipped > 0:
            print(f"Skipping {skipped} disabled applications")
        print(f"Checking {total_apps} enabled applications...")
        print()
        completed = 0

        # Thread-local storage for capturing output per app
        _thread_local = threading.local()
        _real_stdout = sys.stdout
        _print_lock = threading.Lock()

        class ThreadLocalStdout:
            """Redirects writes to a thread-local buffer when available"""

            def write(self, text):
                buf = getattr(_thread_local, "buffer", None)
                if buf is not None:
                    buf.write(text)
                else:
                    _real_stdout.write(text)

            def flush(self):
                _real_stdout.flush()

        def buffered_check(row_num):
            """Run check_single_application with buffered output"""
            _thread_local.buffer = io.StringIO()
            try:
                self.check_single_application(row_num)
                return _thread_local.buffer.getvalue()
            finally:
                _thread_local.buffer = None

        # Redirect stdout to thread-local buffers
        sys.stdout = ThreadLocalStdout()

        # Pre-populate aioesphomeapi singleton cache to avoid cross-loop Future errors
        # when running in ThreadPoolExecutor (each thread gets its own event loop).
        # The @singleton("local_timezone") decorator caches a Future bound to one loop;
        # pre-populating with the resolved value ensures all threads find a plain string.
        try:
            import asyncio
            from aioesphomeapi.singleton import _SINGLETON_CACHE
            async def _prepopulate_timezone():
                from aioesphomeapi.timezone import get_local_timezone
                return await get_local_timezone()
            result = asyncio.run(_prepopulate_timezone())
            _SINGLETON_CACHE["local_timezone"] = result
        except Exception:
            pass  # If aioesphomeapi isn't available, ESPHome checks will handle it

        # Use ThreadPoolExecutor for concurrent checking
        try:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_row = {
                    executor.submit(buffered_check, row_num): row_num
                    for row_num in row_nums
                }

                for future in as_completed(future_to_row):
                    row_num = future_to_row[future]
                    completed += 1
                    try:
                        output = future.result()
                        if output:
                            with _print_lock:
                                _real_stdout.write(output)
                                if not output.endswith("\n"):
                                    _real_stdout.write("\n")
                        if completed % 5 == 0 or completed == total_apps:
                            with _print_lock:
                                _real_stdout.write(
                                    f"Progress: {completed}/{total_apps} applications checked\n"
                                )
                    except Exception as e:
                        with _print_lock:
                            _real_stdout.write(f"Error checking row {row_num}: {e}\n")
        finally:
            sys.stdout = _real_stdout

        self.save_workbook()
        print("=" * 50)
        print(
            f"Version check completed! Checked {total_apps} applications concurrently."
        )

    def show_summary(self):
        """Display a summary of version status (enabled applications only)"""
        if self.worksheet is None:
            print("No worksheet loaded")
            return

        print("\nVersion Summary (Enabled Applications):")
        print("=" * 40)

        # Count statuses for enabled applications only
        status_counts = {}
        total_apps = 0
        disabled_apps = 0

        for row_num in range(2, self.worksheet.max_row + 1):
            # Check if application is enabled
            is_enabled = True
            if "Enabled" in self.columns:
                enabled_cell = self.worksheet[f"{self.columns['Enabled']}{row_num}"]
                is_enabled = enabled_cell.value is True

            if not is_enabled:
                disabled_apps += 1
                continue

            if "Status" in self.columns:
                status_cell = self.worksheet[f"{self.columns['Status']}{row_num}"]
                status = status_cell.value if status_cell.value else "Unknown"
                status_counts[status] = status_counts.get(status, 0) + 1
            total_apps += 1

        for status, count in status_counts.items():
            icon = self.STATUS_ICONS.get(status, "")
            print(f"{icon} {status}: {count}")

        print(f"\nTotal Enabled Applications: {total_apps}")
        if disabled_apps > 0:
            print(f"Disabled Applications: {disabled_apps}")

        # Show outdated applications (enabled only)
        print(f"\n⚠️  Applications needing updates:")
        if all(
            col in self.columns
            for col in [
                "Status",
                "Name",
                "Instance",
                "Current_Version",
                "Latest_Version",
            ]
        ):
            for row_num in range(2, self.worksheet.max_row + 1):
                # Check if application is enabled
                is_enabled = True
                if "Enabled" in self.columns:
                    enabled_cell = self.worksheet[f"{self.columns['Enabled']}{row_num}"]
                    is_enabled = enabled_cell.value is True

                if not is_enabled:
                    continue

                status_cell = self.worksheet[f"{self.columns['Status']}{row_num}"]
                if status_cell.value == "Update Available":
                    name_cell = self.worksheet[f"{self.columns['Name']}{row_num}"]
                    instance_cell = self.worksheet[
                        f"{self.columns['Instance']}{row_num}"
                    ]
                    current_cell = self.worksheet[
                        f"{self.columns['Current_Version']}{row_num}"
                    ]
                    latest_cell = self.worksheet[
                        f"{self.columns['Latest_Version']}{row_num}"
                    ]

                    name = name_cell.value if name_cell.value else ""
                    instance = instance_cell.value if instance_cell.value else ""
                    current = current_cell.value if current_cell.value else "N/A"
                    latest = latest_cell.value if latest_cell.value else "N/A"

                    app_display = f"{name}-{instance}" if instance != "prod" else name
                    print(f"  {app_display}: {current} -> {latest}")
        else:
            print("  Unable to show updates - missing required columns")

    def show_applications(self):
        """Show all applications in a formatted table with dynamic column widths"""
        if self.worksheet is None:
            print("No worksheet loaded")
            return

        # First pass: collect all data and calculate maximum widths
        all_data = []
        max_widths = {
            "index": 4,
            "name": len("Name"),
            "instance": len("Instance"),
            "current": len("Current"),
            "latest": len("Latest"),
            "status": 3,  # Just space for status icon
        }

        for row_num in range(2, self.worksheet.max_row + 1):
            # Skip disabled applications
            if "Enabled" in self.columns:
                enabled_cell = self.worksheet[f"{self.columns['Enabled']}{row_num}"]
                if enabled_cell.value is not True:
                    continue

            row_data = self.get_row_data(row_num)

            name = str(row_data.get("Name", "")) if row_data.get("Name") else ""
            instance = (
                str(row_data.get("Instance", "")) if row_data.get("Instance") else ""
            )
            current = (
                str(row_data.get("Current_Version", ""))
                if row_data.get("Current_Version")
                else ""
            )
            latest = (
                str(row_data.get("Latest_Version", ""))
                if row_data.get("Latest_Version")
                else ""
            )
            status = str(row_data.get("Status", "")) if row_data.get("Status") else ""

            # Update max widths
            max_widths["name"] = max(max_widths["name"], len(name))
            max_widths["instance"] = max(max_widths["instance"], len(instance))
            max_widths["current"] = max(max_widths["current"], len(current))
            max_widths["latest"] = max(max_widths["latest"], len(latest))

            all_data.append(
                {
                    "name": name,
                    "instance": instance,
                    "current": current,
                    "latest": latest,
                    "status": status,
                }
            )

        # Calculate total width and print header
        total_width = (
            sum(max_widths.values()) + len(max_widths) * 2
        )  # +2 for spacing between columns

        print("\nApplications:")
        print("=" * total_width)

        # Print header with dynamic widths
        print(
            f"{'#':<{max_widths['index']}} {'Name':<{max_widths['name']}} {'Instance':<{max_widths['instance']}} {'Current':<{max_widths['current']}} {'Latest':<{max_widths['latest']}} {'':<{max_widths['status']}}"
        )
        print("-" * total_width)

        # Print each row with dynamic widths
        for index, data in enumerate(all_data):
            status_icon = (
                self.STATUS_ICONS.get(data["status"], "") if data["status"] else ""
            )

            print(
                f"{index:<{max_widths['index']}} {data['name']:<{max_widths['name']}} {data['instance']:<{max_widths['instance']}} {data['current']:<{max_widths['current']}} {data['latest']:<{max_widths['latest']}} {status_icon:<{max_widths['status']}}"
            )

        print(f"\nTotal: {len(all_data)} applications")

    def show_updates(self):
        """Show only applications with updates available (enabled only)"""
        if self.worksheet is None:
            print("No worksheet loaded")
            return

        required_cols = [
            "Status",
            "Name",
            "Instance",
            "Current_Version",
            "Latest_Version",
        ]
        if not all(col in self.columns for col in required_cols):
            print("Unable to show updates - missing required columns")
            return

        updates = []
        max_widths = {
            "index": 4,
            "name": len("Name"),
            "instance": len("Instance"),
            "current": len("Current"),
            "latest": len("Latest"),
            "status": 3,
        }

        for row_num in range(2, self.worksheet.max_row + 1):
            if "Enabled" in self.columns:
                enabled_cell = self.worksheet[f"{self.columns['Enabled']}{row_num}"]
                if enabled_cell.value is not True:
                    continue

            status_cell = self.worksheet[f"{self.columns['Status']}{row_num}"]
            if status_cell.value != "Update Available":
                continue

            name_cell = self.worksheet[f"{self.columns['Name']}{row_num}"]
            instance_cell = self.worksheet[f"{self.columns['Instance']}{row_num}"]
            current_cell = self.worksheet[f"{self.columns['Current_Version']}{row_num}"]
            latest_cell = self.worksheet[f"{self.columns['Latest_Version']}{row_num}"]

            name = str(name_cell.value) if name_cell.value else ""
            instance = str(instance_cell.value) if instance_cell.value else ""
            current = str(current_cell.value) if current_cell.value else ""
            latest = str(latest_cell.value) if latest_cell.value else ""

            max_widths["name"] = max(max_widths["name"], len(name))
            max_widths["instance"] = max(max_widths["instance"], len(instance))
            max_widths["current"] = max(max_widths["current"], len(current))
            max_widths["latest"] = max(max_widths["latest"], len(latest))

            updates.append(
                {
                    "name": name,
                    "instance": instance,
                    "current": current,
                    "latest": latest,
                    "status": "Update Available",
                }
            )

        total_width = sum(max_widths.values()) + len(max_widths) * 2

        print("\nApplications Needing Updates:")
        print("=" * total_width)
        print(
            f"{'#':<{max_widths['index']}} {'Name':<{max_widths['name']}} {'Instance':<{max_widths['instance']}} {'Current':<{max_widths['current']}} {'Latest':<{max_widths['latest']}} {'':<{max_widths['status']}}"
        )
        print("-" * total_width)

        for index, data in enumerate(updates):
            status_icon = self.STATUS_ICONS.get(data["status"], "")
            print(
                f"{index:<{max_widths['index']}} {data['name']:<{max_widths['name']}} {data['instance']:<{max_widths['instance']}} {data['current']:<{max_widths['current']}} {data['latest']:<{max_widths['latest']}} {status_icon:<{max_widths['status']}}"
            )

        print(f"\nTotal: {len(updates)} applications")


def main():
    """Main interactive interface"""
    vm = VersionManager()

    while True:
        print("\n" + "=" * 50)
        print("Version Checker Management System")
        print("=" * 50)
        print("1. Check all applications")
        print("2. Check single application")
        print("3. Show summary")
        print("4. List all applications")
        print("5. List updates only")
        print("6. Exit")

        choice = input("\nEnter choice (1-6): ").strip()

        if choice == "1":
            vm.check_all_applications()
        elif choice == "2":
            vm.show_applications()
            try:
                index = int(input("Enter application number (0-based): "))
                row_num = index + 2  # Convert to Excel row number
                if 2 <= row_num <= vm.worksheet.max_row:
                    vm.check_single_application(row_num)
                    vm.save_workbook()
                else:
                    print("Invalid index")
            except ValueError:
                print("Please enter a valid number")
        elif choice == "3":
            vm.show_summary()
        elif choice == "4":
            vm.show_applications()
        elif choice == "5":
            vm.show_updates()
        elif choice == "6":
            break
        else:
            print("Invalid choice. Please enter 1-6.")


if __name__ == "__main__":
    main()
