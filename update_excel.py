#!/usr/bin/env python3

from openpyxl import load_workbook
from config import EXCEL_FILE_PATH

def update_excel_structure():
    """Update existing Excel file structure while preserving all data and formatting"""
    
    excel_path = EXCEL_FILE_PATH
    
    # Load existing workbook to preserve formatting
    try:
        workbook = load_workbook(excel_path)
        worksheet = workbook['Sheet1']
        print(f"Loaded existing workbook with {worksheet.max_row - 1} applications")
    except FileNotFoundError:
        print("ERROR: No existing Excel file found!")
        return None
    except Exception as e:
        print(f"Error reading existing Excel file: {e}")
        return None
    
    # Column mapping for cleanup
    column_mapping = {
        'Name': 'A',
        'Instance': 'B', 
        'Type': 'C',
        'Category': 'D',
        'Target': 'E',
        'GitHub': 'F',
        'DockerHub': 'G',
        'Current_Version': 'H',
        'Latest_Version': 'I',
        'Status': 'J',
        'Last_Checked': 'K',
        'Check_Current': 'L',
        'Check_Latest': 'M'
    }
    
    # Check if any unwanted columns exist and remove them
    # This would need manual identification of unwanted columns by their letter positions
    # For safety, we'll just report the current structure
    
    print("Current column structure:")
    header_row = 1
    for col_name, col_letter in column_mapping.items():
        cell = worksheet[f'{col_letter}{header_row}']
        print(f"  {col_letter}: {cell.value}")
    
    # Auto-adjust column widths while preserving formatting
    for col_name, col_letter in column_mapping.items():
        column = worksheet.column_dimensions[col_letter]
        
        # Calculate max width for this column
        max_length = 0
        for row in worksheet.iter_rows(min_col=ord(col_letter) - ord('A') + 1, 
                                       max_col=ord(col_letter) - ord('A') + 1):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass  # Skip cells with problematic values
        
        # Set adjusted width (with reasonable limits)
        adjusted_width = min(max_length + 2, 50)
        if adjusted_width > column.width:  # Only increase width, never decrease
            column.width = adjusted_width
    
    # Save the updated workbook
    try:
        workbook.save(excel_path)
        print(f"Updated Excel file with formatting preserved: {excel_path}")
        return excel_path
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return None

def remove_unwanted_columns():
    """Remove specific unwanted columns while preserving formatting"""
    excel_path = EXCEL_FILE_PATH
    
    try:
        workbook = load_workbook(excel_path)
        worksheet = workbook['Sheet1']
        print(f"Loaded workbook for column cleanup")
        
        # Define unwanted column names to look for in header row
        unwanted_columns = ['Update_Count', 'Full_Version', 'Update_Size', 'Update_Details']
        
        # Find which columns to delete by checking header row
        columns_to_delete = []
        for col_idx, col in enumerate(worksheet.iter_cols(max_row=1), 1):
            header_cell = col[0]
            if header_cell.value in unwanted_columns:
                col_letter = chr(ord('A') + col_idx - 1)
                columns_to_delete.append((col_idx, col_letter, header_cell.value))
        
        if not columns_to_delete:
            print("No unwanted columns found.")
            return excel_path
        
        # Delete columns in reverse order to maintain indices
        for col_idx, col_letter, col_name in reversed(columns_to_delete):
            print(f"Deleting column {col_letter} ({col_name})")
            worksheet.delete_cols(col_idx)
        
        # Save changes
        workbook.save(excel_path)
        print(f"Removed {len(columns_to_delete)} unwanted columns from Excel file")
        return excel_path
        
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return None

if __name__ == "__main__":
    print("Excel Structure Update Utility (openpyxl-only)")
    print("=" * 50)
    
    choice = input("Choose operation:\n1. Update structure and auto-size columns\n2. Remove unwanted columns\nEnter choice (1-2): ").strip()
    
    if choice == '1':
        update_excel_structure()
    elif choice == '2':
        remove_unwanted_columns()
    else:
        print("Invalid choice. Please enter 1 or 2.")